from openpyxl import load_workbook

from console.excel_find import excel_find
from console.questions.switch import switch


class execl_read:
    # 读取一个excel的所有内容
    @staticmethod
    def read_execl(file_path, page, *book_name):
        # 实例化一个workbook对象
        wb = load_workbook(file_path)
        # 获取一个excel的所有sheet
        wb_sheet = wb.sheetnames
        # 遍历一个excel的所有sheet
        # for page in range(len(wb_sheet)):
        ws = wb[wb_sheet[page]]
        # 这里可以接参数，调用固定的sheet处理函数
        switch[page](ws, *book_name)

    # 获取书名列表
    @staticmethod
    def find_book_name(book_path):
        # BASE_DIR = excel_find.input_path()
        find_excel = excel_find()
        execl_list = find_excel.find_execl(book_path, "*.xlsx")
        book_name_list = []
        # 路径切片，返回书名
        for i in range(len(execl_list)):
            # print(execl_list[i])
            book_name = execl_list[i].split('\\')
            book_name = book_name[-1].split('.')
            # print(book_name)
            book_name_list.append(book_name[0])
            # print(book_name_list)
        return book_name_list

